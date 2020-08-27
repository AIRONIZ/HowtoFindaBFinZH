#coding = utf-8
'''
Author: 上九万里
LastEditors: 上九万里
version: 
Date: 2020-08-26 23:27:11
LastEditTime: 2020-08-27 10:22:33
Description: 将获取到的数据存入excel表格中
'''
import openpyxl as ox
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font

def init_sheet():
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = '答主Id'
    ws['B1'] = '答主主页链接'
    ws['C1'] = '点赞数'
    ws['D1'] = '评论数'
    ws['E1'] = '回答链接'
    ws['F1'] = '回答原文'
    ws['G1'] = '坐标'
    ws['H1'] = '身高'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 80
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8

    wb.save('./result.xlsx')

def save_data(comments):
    wb = ox.load_workbook("./result.xlsx")
    ws = wb['Sheet']

    for comment in comments:
        ws.append(comment)
        wb.save('./result.xlsx')







